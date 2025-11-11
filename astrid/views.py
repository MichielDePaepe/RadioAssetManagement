# astrid/views.py
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views import View
from django.views.generic import TemplateView
from django.views.generic.detail import DetailView
from django.urls import reverse
from django.utils.translation import gettext as _
from django.core.exceptions import PermissionDenied
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import HttpResponseBadRequest, HttpResponseForbidden
from django.db.models import Q


from django.db import transaction
import openpyxl

from helpdesk.models import *
from radio.models import *
from .models import *

import logging
logger = logging.getLogger(__name__)


class UploadSubscriptionsView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'radio.can_upload_subscriptions'
    template_name = "astrid/upload_subscriptions.html"

    def post(self, request):
        try:
            # Get the uploaded Excel file from the request
            excel_file = request.FILES["excelFile"]

            # Load the Excel workbook (read-only, with formulas resolved to values)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            # Extract the header row and map column names to their index
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col = {name: idx for idx, name in enumerate(header)}

            # Perform all DB operations in a single atomic transaction
            with transaction.atomic():
                excel_subs = set()

                # Get existing subscriptions linked to owned customers (TEI, ISSI pairs)
                existing_subs = set(Subscription.objects.filter(issi__customer__owner=True, DMO_only=False).values_list('radio__TEI', 'issi__number'))

                # Iterate through each row in the Excel file
                for row in ws.iter_rows(min_row=2):
                    tei_cell = row[col['TEI']].value
                    issi_cell = row[col['ISSI']].value
                    alias_cell = row[col['CICAlias']].value
                    model_type_cell = row[col['ModelType']].value

                    # Skip rows with missing TEI or ISSI
                    if tei_cell is None or issi_cell is None:
                        continue

                    # Skip spare subscriptions
                    if model_type_cell == "Spare subscription":
                        continue

                    try:
                        # Convert TEI to string and normalize (remove trailing 0 if length is 15)
                        tei_str = str(tei_cell).strip()
                        if len(tei_str) == 15 and tei_str[-1] == "0":
                            tei_str = tei_str[0:-1]
                        tei = int(tei_str)

                        # Parse ISSI number as integer
                        issi_number = int(str(issi_cell).strip())
                    except ValueError:
                        # Invalid TEI or ISSI formatting -> skip row with error message
                        messages.error(request, f"Onjuiste waarde TEI={tei_cell}, ISSI={issi_cell}")
                        continue

                    # Parse alias (optional field)
                    alias = str(alias_cell).strip() if alias_cell else ""

                    try:
                        # Get or create the radio object by TEI
                        radio, _ = Radio.objects.get_or_create(TEI=tei)
                    except ValueError as e:
                        # Invalid TEI -> skip row with error message
                        messages.error(request, f"{e}")
                        continue

                    # Get or create the ISSI object
                    issi, _ = ISSI.objects.get_or_create(number=issi_number)

                    if (tei, issi_number) not in existing_subs:
                        # If ISSI already belongs to another subscription, delete that one
                        other_subscirption_with_issi = Subscription.objects.filter(issi=issi)
                        if other_subscirption_with_issi:
                            other_subscirption_with_issi.delete()

                        # Create a new subscription for this TEI/ISSI combinatiom
                        Subscription.objects.create(
                            radio=radio,
                            issi=issi,
                            astrid_alias=alias
                        )
                        
                    else:
                        # Update alias if subscription already exists but alias changed
                        sub = Subscription.objects.get(radio=radio, issi=issi_number)
                        if sub.astrid_alias != alias:
                            sub.astrid_alias = alias
                            sub.save()

                    # Track the subscriptions present in Excel
                    excel_subs.add((tei, issi_number))

                # Remove subscriptions that no longer exist in the Excel file
                to_delete = existing_subs - excel_subs
                for tei_del, issi_del in to_delete:
                    subscription_to_delete = Subscription.objects.filter(radio__TEI=tei_del, issi__number=issi_del)
                    logger.info(f"Delete subscripton: {subscription_to_delete}")
                    subscription_to_delete.delete()

            # Show success message with summary
            messages.success(request, f"Succesvol verwerkt. {len(excel_subs)} abonnomenten geregistreerd, {len(to_delete)} abonnomenten verwijderd.")

        except KeyError as e:
            # Raised if a required column is missing in the Excel file
            messages.error(request, f"Kolom ontbreekt in Excel: {e}")
        except Exception as e:
            # Catch any other unexpected error, include ISSI/TEI context if available
            messages.error(request, f"Er is een fout opgetreden: {e} (ISSI: {issi_cell}, TEI: {tei_cell})")

        # Return the template with updated context after processing
        return self.get(request)



class VTEIRequestCreateView(TemplateView):
    template_name = "astrid/vtei_request.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        radio_id = self.request.GET.get("radio")
        context["preselected_radio"] = None
        if radio_id:
            try:
                context["preselected_radio"] = Radio.objects.get(pk=radio_id)
            except Radio.DoesNotExist:
                pass
        return context

    def post(self, request):
        try:
            old_radio_pk = request.POST.get("old-radio")
            new_radio_pk = request.POST.get("new-radio")

            if not (old_radio_pk and new_radio_pk):
                raise Exception(_("Both an old and a new radio need to be selected."))

            old_radio = Radio.objects.get(pk=int(old_radio_pk))
            new_radio = Radio.objects.get(pk=int(new_radio_pk))

            if not old_radio.is_active:
                raise Exception(_("The radio selected as old radio has no subscription."))

            if new_radio.is_active:
                raise Exception(_("The radio selected as new radio already has a subscription."))

            # Controleer of er al een open request bestaat
            conflict = Request.objects.filter(
                (Q(radio=old_radio) | Q(old_radio=old_radio) |
                 Q(radio=new_radio) | Q(old_radio=new_radio)),
                ticket_type__code="ASTRID_REQUEST"
            ).exclude(status__code="CLOSED").first()

            if conflict:
                ticket_url = reverse("astrid:request_detail", kwargs={"pk": conflict.pk})
                raise Exception(
                    _("There is already an open Astrid request for one of the radios: "
                      "<a href='{url}'>#{id}</a>").format(url=ticket_url, id=conflict.pk)
                )

            # Omschrijving vooraf invullen met referentie naar ticket
            description = request.POST.get("request_description", "")

            # Nieuwe request aanmaken
            new_request = Request.objects.create(
                request_type=Request.RequestType.VTEI,
                old_radio=old_radio,
                old_issi=old_radio.subscription.issi,
                new_issi=old_radio.subscription.issi,
                radio=new_radio,
                description=description,
            )

            # TicketLog aanmaken als ticket bestaat
            if ticket:
                TicketLog.objects.create(
                    ticket=ticket,
                    user=request.user,
                    status_after=ticket.status,
                    note=f"Astrid VTEI request #{new_request.pk} aangemaakt op basis van dit ticket.",
                )

            messages.success(request, _("VTEI request succesvol aangemaakt"))
            return redirect(reverse("astrid:request_detail", kwargs={"pk": new_request.pk}))

        except Exception as e:
            if settings.DEBUG:
                raise

            messages.error(request, str(e))
            return redirect(request.path)



class ActivationRequestCreateView(TemplateView):
    template_name = "astrid/activation_request.html"

    def post(self, request):
        try:
            # Get the selected radio primary key from the POST data
            radio_pk = request.POST.get("radio")

            # A radio must be selected, otherwise raise an exception
            if not radio_pk:
                raise(Exception(_("A radio needs to be selected.")))

            # Fetch the radio object from the database
            radio = Radio.objects.get(pk=int(radio_pk))
                
            # Prevent activation if the radio already has a subscription
            if radio.is_active:
                url = reverse("radio:detail", kwargs={"pk": radio.pk})
                raise(Exception(_("The selected <a href='{url}'>radio</a> already has a subscription.").format(url=url)))

            # Get the ISSI (Individual Short Subscriber Identity) value from POST data
            issi_val = request.POST.get("issi")

            # A valid ISSI must be provided
            if not issi_val:
                raise(Exception(_("A ISSI needs to be provided.")))

            # Fetch the ISSI object from the database
            issi = ISSI.objects.get(pk=int(issi_val))

            # Prevent activation if this ISSI is already linked to a subscription
            if hasattr(issi, "subscription"):
                url = reverse("radio:detail", kwargs={"pk": issi.subscription.radio.pk})
                raise(Exception(_("The ISSI <a href='{url}'>{issi}</a> is already activated.").format(url=url, issi=issi_val)))

            # Check if there is already an open ASTRID request ticket linked to this radio
            req = Request.objects.filter((Q(radio=radio) | Q(old_radio=radio)) & Q(ticket_type__code="ASTRID_REQUEST")).exclude(status__code = "CLOSED").first()
            if req:
                ticket_url = reverse("astrid:request_detail", kwargs={"pk": req.pk})
                radio_url = reverse("radio:detail", kwargs={"pk": radio.pk})
                raise(Exception(_("The <a href='{radio_url}'>radio</a> has an open request ticket: <a href='{ticket_url}'>#{ticket_id}</a>").format(radio_url=radio_url, ticket_url=ticket_url, ticket_id=req.pk)))

            # Check if there is already an open ASTRID request ticket linked to this ISSI
            req = Request.objects.filter((Q(old_issi=issi) | Q(new_issi=issi)) & Q(ticket_type__code="ASTRID_REQUEST")).exclude(status__code = "CLOSED").first()
            if req:
                ticket_url = reverse("astrid:request_detail", kwargs={"pk": req.pk})
                raise(Exception(_("The ISSI is already used in an open request ticket: <a href='{ticket_url}'>#{ticket_id}</a>").format(ticket_url=ticket_url, ticket_id=req.pk)))

            # If no conflicts are found, create a new activation request
            Request.objects.create(
                request_type = Request.RequestType.ACTIVATION,
                new_issi = issi,
                radio = radio,
                description = request.POST.get("request_description"),
            )

            # Show a success message to the user
            messages.success(request, _("Activation request successful"))

        except Exception as e:
            # Catch all raised exceptions and show them as error messages
            messages.error(request, str(e))

        # Redirect back to the same page after handling the request
        return redirect(request.path)



class RequestOverviewView(TemplateView):
    template_name = "astrid/request_overview.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["requests"] = (
            Request.objects
            .filter(ticket_type__code="ASTRID_REQUEST")
            .exclude(status__code="CLOSED")
            .order_by("-id")
        )
        return context



class RequestDetailView(DetailView):
    model = Request
    template_name = "astrid/request_detail.html"

    def post(self, request, *args, **kwargs):
        obj = self.get_object()

        request_type = request.POST.get("type")
        action = request.POST.get("action")
        note = request.POST.get("note")

        try:

            if request_type == "astrid_request_submitted":
                if not request.user.has_perm("astrid.has_access_to_myastrid"):
                    raise PermissionDenied

                if action == "request_submitted":
                    astrid_ticket = request.POST.get("astrid_ticket")

                    if not astrid_ticket:
                        messages.error(request, _("Geen astrid ticket nummer opgegeven"))
                        return redirect(request.path)

                    obj.external_reference = astrid_ticket
                    obj.save() 
                    obj.start_execution(user=request.user, note=note)

                elif action == "refused":
                    if not note:
                        messages.error(request, _("Geef een reden op waarom de aanvraag geweigerd werd"))
                        return redirect(request.path)
                    obj.mark_closed(user=request.user, note=note)
            
            elif request_type == "feedback_from_astrid":
                if not request.user.has_perm("astrid.has_access_to_myastrid"):
                    raise PermissionDenied

                if action == "precessed":
                    obj.mark_waiting_verification(user=request.user, note=note)
                elif action == "refused":
                    obj.mark_closed(user=request.user, note=note or _("Request refused by astrid"))

            elif request_type == "validate_activation":
                if not request.user.has_perm("astrid.can_verify_requests"):
                    raise PermissionDenied

                if not note:
                    messages.error(request, _("Opmerking mag niet leeg zijn"))
                    return redirect(request.path)

                if action == "radio_is_working":
                    obj.mark_verified(user=request.user, note=note)

                if action == "radio_is_not_working":
                    obj.add_log(user=request.user, note=note)
            
            else:
                return HttpResponseBadRequest("Invalid POST")
        
        except Exception as e:
            if settings.DEBUG:
                raise
            
            messages.error(request, str(e))
            return redirect(request.path)

        return redirect("astrid:request_overview")





