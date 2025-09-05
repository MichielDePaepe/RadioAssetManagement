# python manage.py shell
# from import_subscriptions import import_subscriptions
# import_subscriptions("report.xlsx")


import openpyxl
from radio.models import Radio, ISSI, Subscription
from django.db import transaction

def set_models(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(header)}

    with transaction.atomic():
        excel_subs = set()
        existing_subs = set(Subscription.objects.values_list('radio__TEI', 'issi__number'))

        for row in ws.iter_rows(min_row=2):
            tei_cell = row[col['TEI']].value
            issi_cell = row[col['ISSI']].value
            alias_cell = row[col['CICAlias']].value

            if tei_cell is None or issi_cell is None:
                continue

            tei = int(str(tei_cell).strip())
            issi_number = int(str(issi_cell).strip())
            alias = str(alias_cell).strip() if alias_cell else ""

            radio, _ = Radio.objects.get_or_create(TEI=tei)
            issi, _ = ISSI.objects.get_or_create(number=issi_number)

            if (tei, issi_number) not in existing_subs:
                Subscription.objects.create(
                    radio=radio,
                    issi=issi,
                    astrid_alias=alias
                )
            else:
                sub = Subscription.objects.get(radio=radio, issi=issi_number)
                if sub.astrid_alias != alias:
                    sub.astrid_alias = alias
                    sub.save()

            excel_subs.add((tei, issi_number))

        # Verwijderen wat niet in Excel zit
        to_delete = existing_subs - excel_subs
        for tei_del, issi_del in to_delete:
            Subscription.objects.filter(radio__TEI=tei_del, issi__number=issi_del).delete()

    print("Import en sync klaar.")
